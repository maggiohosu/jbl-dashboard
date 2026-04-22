# -*- coding: utf-8 -*-
"""
JBL 판매 대시보드 — 백엔드 데이터 처리 엔진
Excel 데이터를 읽어 dash_data.json을 생성합니다.

실행: python data_engine.py
"""

import os
import re
import json
from datetime import date, datetime, timedelta
from collections import defaultdict

import pandas as pd

# ═══════════════════════════════════════════════════════════
#  파일 경로 설정
# ═══════════════════════════════════════════════════════════
FOLDER         = r"C:\Users\SG\Desktop\프로그램\판매데이터"
MASTER_FILE    = os.path.join(FOLDER, "데이터_마스터.xlsx")
TARGET_FILE    = os.path.join(FOLDER, "월별목표.xlsx")
SKU_MODEL_FILE = os.path.join(FOLDER, "SKU 및 모델명.xlsx")
OUTPUT_JSON    = os.path.join(FOLDER, "dash_data.json")

# ═══════════════════════════════════════════════════════════
#  마켓 분류 설정
# ═══════════════════════════════════════════════════════════
MARKET_KEYWORDS = {
    "네이버":  "네이버",
    "쿠팡":    "쿠팡",
    "알리":    "알리",
    "SSG":     "SSG",
    "신세계":  "SSG",
    "지마켓":  "지마켓/옥션",
    "옥션":    "지마켓/옥션",
    "11번가":  "11번가",
}
OTHER_MALL = {
    "29CM", "CJ온스타일", "W컨셉", "마켓컬리", "무신사", "아트박스",
    "오늘의집", "카카오-선물", "카카오-쇼핑", "크림", "톡스토어", "현대쇼핑",
    "더현대닷컴", "롯데쇼핑", "롯데온", "디플롯", "토스", "컴퓨존", "퍼스트몰", "자사몰",
}
MARKET_ORDER = ["네이버", "쿠팡", "알리", "SSG", "지마켓/옥션", "11번가", "기타몰", "도소매"]

# ═══════════════════════════════════════════════════════════
#  카테고리 분류 설정 (process_sales.py와 동일)
# ═══════════════════════════════════════════════════════════
CATEGORY_MAP = {
    "이어폰":     "TWS",
    "무선":       "헤드셋",
    "무선(NC)":   "헤드셋",
    "스피커":     "스피커",
    "하이파이":   "럭셔리",
    "북쉘프":     "럭셔리",
    "턴테이블":   "럭셔리",
    "유선(키즈)": "헤드셋",
    "무선(키즈)": "헤드셋",
    "+우퍼":      "사운드바",
    "올인원":     "사운드바",
    "G.헤드셋":   "QUANTUM",
    "G.이어폰":   "QUANTUM",
    "PC마이크":   "QUANTUM",
    "배터리":     "액세서리",
    "스탠드":     "액세서리",
    "L75MS":      "액세서리",
    "TPRO3":      "액세서리",
}
CATEGORY_ORDER = ["TWS", "헤드셋", "스피커", "럭셔리", "사운드바", "QUANTUM", "액세서리"]

# 모델별 카테고리 강제 지정 (원가표 오류 보정용)
MODEL_CATEGORY_OVERRIDE = {
    "JBLXTREME4BLKAS": "스피커",
}

# SKU 별칭 매핑 (판매 데이터 코드 ↔ 원가표 코드 불일치 보정)
SKU_ALIASES = {
    "JBLTOM3":        "JBLTOURONEM3",
    "TOURPRO3 케이스": "TPRO3 케이스",
}

# ── 원가표에 없는 제품 직접 추가 ────────────────────────────
EXTRA_PRODUCTS = {
    "JBLBAR300PRO":   {"category": "사운드바", "price": 0, "model_group": "JBL BAR 300PRO",          "brand": "JBL"},
    "JBLBAR21DBM2":   {"category": "사운드바", "price": 0, "model_group": "JBL BAR 2.1 DBM2",        "brand": "JBL"},
    "JBLJR310BT":     {"category": "헤드셋",   "price": 0, "model_group": "JBL JR310BT",             "brand": "JBL"},
    "JBLJR310":       {"category": "헤드셋",   "price": 0, "model_group": "JBL JR310",               "brand": "JBL"},
    "JBLSNDGEARSNS":  {"category": "헤드셋",   "price": 0, "model_group": "JBL SOUND GEAR SENSE",    "brand": "JBL"},
    "JBLQUANTUMTWS":  {"category": "QUANTUM",  "price": 0, "model_group": "JBL QUANTUM TWS",         "brand": "JBL"},
    "JBLQTUM250BLK":  {"category": "QUANTUM",  "price": 0, "model_group": "JBL QUANTUM 250",         "brand": "JBL"},
    "JBLTOURONEM2":   {"category": "럭셔리",   "price": 0, "model_group": "JBL TOUR ONE M2",         "brand": "JBL"},
    "JBLTOURPRO2":    {"category": "TWS",      "price": 0, "model_group": "JBL TOUR PRO 2",          "brand": "JBL"},
    "JBLWIND3S":      {"category": "스피커",   "price": 0, "model_group": "JBL WIND3S",              "brand": "JBL"},
    "JBLL75MSBGAS":   {"category": "액세서리", "price": 0, "model_group": "L75MS-BG",                "brand": "JBL"},
    "TOURONEM3 이어패드":             {"category": "액세서리", "price": 0, "model_group": "TOURONEM3 이어패드",          "brand": "JBL"},
    "오롤리데이 HAPPIER 패키지":      {"category": "액세서리", "price": 0, "model_group": "오롤리데이 HAPPIER 패키지",  "brand": "JBL"},
    "오롤리데이 컵세트":              {"category": "액세서리", "price": 0, "model_group": "오롤리데이 컵세트",          "brand": "JBL"},
    "단가정산":       {"category": "액세서리", "price": 0, "model_group": "단가정산",                 "brand": "JBL"},
}

# ═══════════════════════════════════════════════════════════
#  색상 설정
# ═══════════════════════════════════════════════════════════
MKT_COLORS = {
    "네이버":     "#0F9D4C",
    "쿠팡":       "#C0392B",
    "알리":       "#E65100",
    "SSG":        "#7B1FA2",
    "지마켓/옥션":"#E65100",
    "11번가":     "#8B0000",
    "기타몰":     "#4527A0",
    "도소매":     "#37474F",
}
CAT_COLORS = {
    "TWS":      "#0071E3",
    "헤드셋":   "#1B8A3B",
    "스피커":   "#E65100",
    "럭셔리":   "#8E24AA",
    "사운드바": "#B71C1C",
    "QUANTUM":  "#212121",
    "액세서리": "#5D6D7E",
}

# ═══════════════════════════════════════════════════════════
#  월별 기본 목표 (단위: 원) — 월별목표.xlsx 없을 때 사용
# ═══════════════════════════════════════════════════════════
DEFAULT_MONTHLY_TARGETS = {
    "네이버":     {1:500000000,2:600000000,3:700000000,4:800000000,
                   5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},
    "쿠팡":       {1:300000000,2:400000000,3:500000000,4:550000000,
                   5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},
    "알리":       {1:20000000,2:22000000,3:24000000,4:25000000,
                   5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},
    "SSG":        {1:80000000,2:100000000,3:110000000,4:120000000,
                   5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},
    "지마켓/옥션":{1:60000000,2:80000000,3:85000000,4:90000000,
                   5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},
    "11번가":     {1:20000000,2:25000000,3:28000000,4:30000000,
                   5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},
    "기타몰":     {1:100000000,2:120000000,3:130000000,4:140000000,
                   5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},
    "도소매":     {1:30000000,2:38000000,3:42000000,4:45000000,
                   5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0},
}


# ═══════════════════════════════════════════════════════════
#  마켓 분류 함수
# ═══════════════════════════════════════════════════════════
def classify_market(raw_name):
    """거래처명 → 마켓명 분류"""
    raw = str(raw_name).strip()
    # (주)아트박스는 도소매로 분류
    if raw == "(주)아트박스" or re.sub(r"\s", "", raw) == "(주)아트박스":
        return "도소매"
    clean = re.sub(r"\(.*?\)", "", raw).strip()
    for key, market in MARKET_KEYWORDS.items():
        if key in clean:
            return market
    if clean in OTHER_MALL:
        return "기타몰"
    for mall in OTHER_MALL:
        if mall in clean:
            return "기타몰"
    if "아트박스" in raw:
        return "기타몰"
    return "도소매"


# ═══════════════════════════════════════════════════════════
#  SKU → 모델명 매핑 로드 (SKU 및 모델명.xlsx)
# ═══════════════════════════════════════════════════════════
def get_sku_model_map():
    """
    SKU 및 모델명.xlsx에서 {SKU명: 모델명} 딕셔너리 반환
    """
    sku_model = {}
    if not os.path.exists(SKU_MODEL_FILE):
        return sku_model
    try:
        df = pd.read_excel(SKU_MODEL_FILE, sheet_name=0, header=0)
        df.columns = [str(c).strip() for c in df.columns]
        for _, row in df.iterrows():
            sku_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            model_val = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if sku_val and model_val and sku_val != "nan" and model_val != "nan":
                sku_model[sku_val] = model_val
        print(f"  [SKU 모델명] {len(sku_model)}개 로드")
    except Exception as e:
        print(f"  [경고] SKU 모델명 로드 실패: {e}")
    return sku_model


# ═══════════════════════════════════════════════════════════
#  원가표 기반 SKU 맵 구축
#  (원가표 없을 경우 EXTRA_PRODUCTS + prefix 매칭으로 폴백)
# ═══════════════════════════════════════════════════════════
def _build_sku_map_from_pricelist():
    """원가표 시트에서 SKU → {category, price, model_group, brand} 로드"""
    sku_map = {}
    # 원가표는 현황 엑셀 파일에 포함됨 — 폴더에서 탐색
    import glob as _glob
    candidates = _glob.glob(os.path.join(FOLDER, "*.xlsx"))
    src = None
    for path in candidates:
        if "마스터" in path or "목표" in path or "SKU" in path:
            continue
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if "원가표" in wb.sheetnames:
                wb.close()
                src = path
                break
            wb.close()
        except Exception:
            pass
    if not src:
        return sku_map
    try:
        import openpyxl
        wb = openpyxl.load_workbook(src, data_only=True)
        ws = wb["원가표"]
        for row in ws.iter_rows(min_row=9, values_only=True):
            brand, prod_type, model_group = row[1], row[2], row[3]
            model_code, price = row[5], row[7]
            if not model_code or not prod_type:
                continue
            base = str(model_code).replace("*", "").strip()
            category = MODEL_CATEGORY_OVERRIDE.get(base) or CATEGORY_MAP.get(str(prod_type).strip())
            if not category:
                continue
            if not isinstance(price, (int, float)):
                if category == "액세서리":
                    price = 0
                else:
                    continue
            sku_map[base] = {
                "category":    category,
                "price":       int(price),
                "model_group": str(model_group) if model_group else base,
                "brand":       str(brand) if brand else "",
            }
        wb.close()
        print(f"  [원가표] {len(sku_map)}개 SKU 로드")
    except Exception as e:
        print(f"  [경고] 원가표 로드 실패: {e}")
    return sku_map


def _build_full_sku_map():
    """원가표 + EXTRA_PRODUCTS + SKU 모델명 파일 통합 SKU 맵 구축"""
    sku_map = _build_sku_map_from_pricelist()
    # EXTRA_PRODUCTS 추가 (원가표에 없는 제품)
    for base, info in EXTRA_PRODUCTS.items():
        if base not in sku_map:
            sku_map[base] = info

    # SKU 및 모델명.xlsx의 모델명으로 model_group 업데이트
    sku_model = get_sku_model_map()
    if sku_model:
        base_to_model = {}
        for sku_code, model_name in sku_model.items():
            best_base, best_len = None, 0
            for base in sku_map:
                if sku_code.startswith(base) and len(base) > best_len:
                    best_base = base
                    best_len = len(base)
            if best_base and best_base not in base_to_model:
                base_to_model[best_base] = model_name
        for base, model_name in base_to_model.items():
            sku_map[base]["model_group"] = model_name
        print(f"  [모델명 반영] {len(base_to_model)}개")
    return sku_map


def _match_sku(sku_code, sku_map):
    """SKU 코드 → sku_map 항목 반환 (없으면 None)"""
    sku = str(sku_code).strip()
    if sku in sku_map:
        return sku_map[sku], sku
    # 별칭 확인
    for alias_prefix, canonical in SKU_ALIASES.items():
        if sku.startswith(alias_prefix) and canonical in sku_map:
            return sku_map[canonical], canonical
    # 가장 긴 prefix 매칭
    best, best_base, best_len = None, None, 0
    for base, info in sku_map.items():
        if sku.startswith(base) and len(base) > best_len:
            best, best_base, best_len = info, base, len(base)
    return best, best_base


# ═══════════════════════════════════════════════════════════
#  카테고리 분류 함수 (SKU prefix 기반 폴백)
# ═══════════════════════════════════════════════════════════
def get_category(sku, sku_map):
    """SKU → 카테고리. sku_map에 있으면 그 값, 없으면 prefix로 추정"""
    info, _ = _match_sku(sku, sku_map)
    if info:
        return info.get("category", "액세서리")
    # prefix 기반 추정
    s = str(sku).upper()
    if s.startswith("JBLQ") or s.startswith("AKGQ"):
        return "QUANTUM"
    if "BAR" in s:
        return "사운드바"
    if any(k in s for k in ["CHARGE", "GO", "FLIP", "WIND", "CLIP", "XTREME", "BOOMBOX"]):
        return "스피커"
    if "TUNE" in s or "LIVE" in s or "VIBE" in s:
        return "TWS"
    if "TOUR ONE" in s or "TOURPRO" in s:
        return "TWS"
    return "액세서리"


# ═══════════════════════════════════════════════════════════
#  마스터 데이터 로드 및 전처리
# ═══════════════════════════════════════════════════════════
def load_master_data():
    """
    데이터_마스터.xlsx 로드 + 기본 전처리
    반환: DataFrame (columns 정규화, 날짜 파싱 포함)
    """
    if not os.path.exists(MASTER_FILE):
        print(f"  [오류] 마스터 파일 없음: {MASTER_FILE}")
        return pd.DataFrame()
    try:
        df = pd.read_excel(MASTER_FILE, sheet_name=0, header=0)
        df.columns = [str(c).strip() for c in df.columns]
        print(f"  [마스터] {len(df)}행 로드")
        return df
    except Exception as e:
        print(f"  [마스터 로드 오류] {e}")
        return pd.DataFrame()


def _parse_order_date(id_no_str):
    """
    일자-No. 파싱 → date 객체
    형식: "26/01/16-533" → 2026-01-16
    """
    try:
        date_part = str(id_no_str).split("-")[0]
        yy, mm, dd = date_part.split("/")
        return date(2000 + int(yy), int(mm), int(dd))
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════
#  보고 연월 자동 감지
# ═══════════════════════════════════════════════════════════
def detect_report_year_month(df):
    """
    마스터 데이터에서 최신 연월(년, 월)을 자동 감지
    반환: (year, month)
    """
    latest = None
    for id_no in df.get("일자-No.", pd.Series(dtype=str)):
        d = _parse_order_date(id_no)
        if d and (latest is None or d > latest):
            latest = d
    if latest:
        return latest.year, latest.month
    # 폴백: 현재 날짜
    now = datetime.now()
    return now.year, now.month


# ═══════════════════════════════════════════════════════════
#  주차 정보 계산
# ═══════════════════════════════════════════════════════════
def get_week_info(year, month):
    """
    보고월 기준 주차 정보 계산
    반환:
      - report_dates: 해당 연도 1월1일 ~ 보고월 말일 date 리스트
      - weeks: {date: 주차번호} (보고월 이전은 0, 보고월부터 1~N)
      - week_groups: {주차번호: [date, ...]}
      - week_labels: {"1": "4/1~4/6", ...}
      - max_week: 최대 주차 수
    """
    first = date(year, month, 1)
    # 보고월 1일이 속한 월요일 = 주1 시작
    week1_start = first - timedelta(days=first.weekday())
    # 보고월 마지막 날
    if month == 12:
        last = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(year, month + 1, 1) - timedelta(days=1)

    # 1월1일 ~ 보고월 말일 전체 날짜
    year_start = date(year, 1, 1)
    report_dates = []
    d = year_start
    while d <= last:
        report_dates.append(d)
        d += timedelta(days=1)

    # 주차 매핑
    weeks = {}
    for d in report_dates:
        if d < week1_start:
            weeks[d] = 0
        else:
            weeks[d] = (d - week1_start).days // 7 + 1

    # 주차 그룹 (보고월 내 날짜만)
    week_groups = defaultdict(list)
    for d in report_dates:
        wk = weeks[d]
        if wk > 0:
            week_groups[wk].append(d)

    # 주차 라벨
    week_labels = {}
    for wk, wk_dates in week_groups.items():
        sd = sorted(wk_dates)
        week_labels[str(wk)] = f"{sd[0].strftime('%m/%d')}~{sd[-1].strftime('%m/%d')}"

    max_week = max(weeks.values()) if weeks else 5

    return report_dates, weeks, dict(week_groups), week_labels, max_week


# ═══════════════════════════════════════════════════════════
#  월별 목표 로드
# ═══════════════════════════════════════════════════════════
def get_monthly_targets():
    """
    월별목표.xlsx → {마켓: {월str: 원}} 반환
    없으면 DEFAULT_MONTHLY_TARGETS 사용
    """
    result = {mkt: {} for mkt in MARKET_ORDER}

    if os.path.exists(TARGET_FILE):
        try:
            # header=1 로 읽으면: 0행=제목, 1행=헤더(마켓,1월,2월...), 2행~=데이터
            df = pd.read_excel(TARGET_FILE, header=1)
            df.columns = [str(c).strip() for c in df.columns]
            # 첫 번째 컬럼이 마켓명
            first_col = df.columns[0]
            for _, row in df.iterrows():
                mkt_name = str(row[first_col]).strip()
                if mkt_name not in MARKET_ORDER:
                    continue
                for mo in range(1, 13):
                    # 컬럼명: "1월", "2월" ... 또는 숫자
                    col_candidates = [f"{mo}월", str(mo), mo]
                    val = 0
                    for cc in col_candidates:
                        if cc in df.columns:
                            raw = row.get(cc, 0)
                            try:
                                val = int(float(raw or 0)) * 10000
                            except Exception:
                                val = 0
                            break
                    result[mkt_name][str(mo)] = val
            print(f"  [월별목표] 로드 완료")
            return result
        except Exception as e:
            print(f"  [경고] 월별목표 로드 실패: {e}")

    # 폴백: 기본값
    for mkt in MARKET_ORDER:
        for mo in range(1, 13):
            result[mkt][str(mo)] = DEFAULT_MONTHLY_TARGETS.get(mkt, {}).get(mo, 0)
    return result


# ═══════════════════════════════════════════════════════════
#  집계 함수
# ═══════════════════════════════════════════════════════════
def _aggregate_data(df, sku_map, report_dates):
    """
    마스터 DataFrame을 집계하여 각종 딕셔너리 반환

    반환 딕셔너리:
      mkt_day_rev   : {마켓: {날짜str: 매출}}
      mkt_day_qty   : {마켓: {날짜str: 수량}}
      cat_day_rev   : {카테고리: {날짜str: 매출}}
      cat_day_qty   : {카테고리: {날짜str: 수량}}
      model_day_rev : {모델명: {날짜str: 매출}}
      model_day_qty : {모델명: {날짜str: 수량}}
      mkt_model_rev : {마켓: {모델명: {날짜str: 매출}}}
      return_rows   : 반품 행 목록 (수량<0)
    """
    report_date_set = set(report_dates)

    mkt_day_rev   = defaultdict(lambda: defaultdict(int))
    mkt_day_qty   = defaultdict(lambda: defaultdict(int))
    cat_day_rev   = defaultdict(lambda: defaultdict(int))
    cat_day_qty   = defaultdict(lambda: defaultdict(int))
    model_day_rev = defaultdict(lambda: defaultdict(int))
    model_day_qty = defaultdict(lambda: defaultdict(int))
    mkt_model_rev = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    mkt_model_qty = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    cat_model_rev = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    # 반품 집계
    ret_mkt_rev   = defaultdict(int)
    ret_mkt_qty   = defaultdict(int)
    ret_mkt_month = defaultdict(lambda: defaultdict(int))  # {마켓: {ym: 반품매출}}
    unknown_skus  = set()

    for _, row in df.iterrows():
        id_no = str(row.get("일자-No.", "")).strip()
        if not id_no or id_no == "nan":
            continue
        order_date = _parse_order_date(id_no)
        if order_date is None or order_date not in report_date_set:
            continue

        sku_raw = str(row.get("품목명(규격)", "")).strip()
        info, sku_base = _match_sku(sku_raw, sku_map)
        if info is None:
            unknown_skus.add(sku_raw)
            continue

        try:
            qty = int(row.get("수량", 0))
        except Exception:
            qty = 0
        if qty == 0:
            continue

        try:
            sales_amount = int(float(str(row.get("판매금액합계", 0)).replace(",", "") or 0))
        except Exception:
            sales_amount = 0

        market    = classify_market(row.get("거래처명", ""))
        cat       = info["category"]
        model_grp = info["model_group"]
        ds        = order_date.isoformat()
        ym        = order_date.strftime("%Y-%m")

        # 반품 처리 (수량 < 0)
        if qty < 0:
            ret_mkt_rev[market]         += abs(sales_amount)
            ret_mkt_qty[market]         += abs(qty)
            ret_mkt_month[market][ym]   += abs(sales_amount)
            continue  # 반품은 매출 집계에서 제외

        # 정상 판매 집계
        mkt_day_rev[market][ds]            += sales_amount
        mkt_day_qty[market][ds]            += qty
        cat_day_rev[cat][ds]               += sales_amount
        cat_day_qty[cat][ds]               += qty
        model_day_rev[model_grp][ds]       += sales_amount
        model_day_qty[model_grp][ds]       += qty
        mkt_model_rev[market][model_grp][ds] += sales_amount
        mkt_model_qty[market][model_grp][ds] += qty
        cat_model_rev[cat][model_grp][ds]    += sales_amount

    if unknown_skus:
        print(f"  [미매핑 SKU {len(unknown_skus)}개]: {sorted(unknown_skus)[:5]}...")

    return {
        "mkt_day_rev":   {m: dict(v) for m, v in mkt_day_rev.items()},
        "mkt_day_qty":   {m: dict(v) for m, v in mkt_day_qty.items()},
        "cat_day_rev":   {c: dict(v) for c, v in cat_day_rev.items()},
        "cat_day_qty":   {c: dict(v) for c, v in cat_day_qty.items()},
        "model_day_rev": {m: dict(v) for m, v in model_day_rev.items()},
        "model_day_qty": {m: dict(v) for m, v in model_day_qty.items()},
        "mkt_model_rev": {mk: {mg: dict(dd) for mg, dd in v.items()} for mk, v in mkt_model_rev.items()},
        "mkt_model_qty": {mk: {mg: dict(dd) for mg, dd in v.items()} for mk, v in mkt_model_qty.items()},
        "cat_model_rev": {c: {mg: dict(dd) for mg, dd in v.items()} for c, v in cat_model_rev.items()},
        "ret_mkt_rev":   dict(ret_mkt_rev),
        "ret_mkt_qty":   dict(ret_mkt_qty),
        "ret_mkt_month": {m: dict(v) for m, v in ret_mkt_month.items()},
    }


# ═══════════════════════════════════════════════════════════
#  MoM (전월 대비) 계산
# ═══════════════════════════════════════════════════════════
def calc_mom(month_data):
    """
    month_data: {dim: {ym_str: value}}
    반환:       {dim: {ym_str: rate}} — ym이 첫 달이면 0.0
    """
    result = {}
    for dim, ym_dict in month_data.items():
        yms = sorted(ym_dict.keys())
        rates = {}
        for i, ym in enumerate(yms):
            if i == 0:
                rates[ym] = 0.0
            else:
                prev_val = ym_dict.get(yms[i - 1], 0)
                curr_val = ym_dict.get(ym, 0)
                if prev_val and prev_val > 0:
                    rates[ym] = round((curr_val - prev_val) / prev_val, 4)
                else:
                    rates[ym] = 0.0
        result[dim] = rates
    return result


# ═══════════════════════════════════════════════════════════
#  WoW (전주 대비) 계산
# ═══════════════════════════════════════════════════════════
def calc_wow(week_data):
    """
    week_data: {dim: {wk_str: value}}
    반환:      {dim: {wk_str: rate}}
    """
    result = {}
    for dim, wk_dict in week_data.items():
        wks = sorted(wk_dict.keys(), key=lambda x: int(x))
        rates = {}
        for i, wk in enumerate(wks):
            if i == 0:
                rates[wk] = 0.0
            else:
                prev_val = wk_dict.get(wks[i - 1], 0)
                curr_val = wk_dict.get(wk, 0)
                if prev_val and prev_val > 0:
                    rates[wk] = round((curr_val - prev_val) / prev_val, 4)
                else:
                    rates[wk] = 0.0
        result[dim] = rates
    return result


# ═══════════════════════════════════════════════════════════
#  TOP3 계산
# ═══════════════════════════════════════════════════════════
def calc_top3(month_data, qty_data=None, top_n=3):
    """
    month_data: {dim: {ym: rev}}
    반환: {ym: [{rank, model/cat, rev, qty}]}
    """
    # ym 목록 수집
    all_yms = set()
    for v in month_data.values():
        all_yms.update(v.keys())

    result = {}
    for ym in sorted(all_yms):
        entries = []
        for dim, ym_dict in month_data.items():
            rev = ym_dict.get(ym, 0)
            qty = (qty_data or {}).get(dim, {}).get(ym, 0) if qty_data else 0
            entries.append({"name": dim, "rev": rev, "qty": qty})
        entries.sort(key=lambda x: x["rev"], reverse=True)
        ranked = []
        for rank, e in enumerate(entries[:top_n], 1):
            item = {"rank": rank, "rev": e["rev"], "qty": e["qty"]}
            # 키 이름은 model/cat에 따라 호출처에서 처리
            item["name"] = e["name"]
            ranked.append(item)
        result[ym] = ranked
    return result


def calc_top3_week(week_data, qty_data=None, top_n=3):
    """week_data: {dim: {wk_str: rev}} → {wk_str: [{rank, name, rev, qty}]}"""
    all_wks = set()
    for v in week_data.values():
        all_wks.update(v.keys())

    result = {}
    for wk in sorted(all_wks, key=lambda x: int(x)):
        entries = []
        for dim, wk_dict in week_data.items():
            rev = wk_dict.get(wk, 0)
            qty = (qty_data or {}).get(dim, {}).get(wk, 0) if qty_data else 0
            entries.append({"name": dim, "rev": rev, "qty": qty})
        entries.sort(key=lambda x: x["rev"], reverse=True)
        ranked = []
        for rank, e in enumerate(entries[:top_n], 1):
            ranked.append({"rank": rank, "name": e["name"], "rev": e["rev"], "qty": e["qty"]})
        result[wk] = ranked
    return result


# ═══════════════════════════════════════════════════════════
#  성장 / 하락 목록 계산
# ═══════════════════════════════════════════════════════════
def calc_growth_decline(month_data, n=10):
    """
    month_data: {dim: {ym: value}}
    n: 상위 n개
    반환: (growth_dict, decline_dict) — 각각 {ym: [{name, prevRev, currRev, growthRate}]}
    """
    all_yms = set()
    for v in month_data.values():
        all_yms.update(v.keys())
    yms_sorted = sorted(all_yms)

    growth  = {}
    decline = {}
    for i, ym in enumerate(yms_sorted):
        if i == 0:
            continue
        prev_ym = yms_sorted[i - 1]
        entries = []
        for dim, ym_dict in month_data.items():
            prev_rev = ym_dict.get(prev_ym, 0)
            curr_rev = ym_dict.get(ym, 0)
            if prev_rev <= 0 and curr_rev <= 0:
                continue
            if prev_rev > 0:
                rate = round((curr_rev - prev_rev) / prev_rev, 4)
            else:
                rate = 1.0 if curr_rev > 0 else 0.0
            entries.append({"name": dim, "prevRev": prev_rev, "currRev": curr_rev, "growthRate": rate})

        up   = sorted([e for e in entries if e["growthRate"] > 0], key=lambda x: x["growthRate"], reverse=True)
        down = sorted([e for e in entries if e["growthRate"] < 0], key=lambda x: x["growthRate"])
        growth[ym]  = up[:n]
        decline[ym] = down[:n]

    return growth, decline


# ═══════════════════════════════════════════════════════════
#  비중(점유율) 계산
# ═══════════════════════════════════════════════════════════
def calc_share(month_data):
    """
    month_data: {dim: {ym: value}}
    반환: {ym: {dim: share_rate}}
    """
    all_yms = set()
    for v in month_data.values():
        all_yms.update(v.keys())

    result = {}
    for ym in sorted(all_yms):
        total = sum(v.get(ym, 0) for v in month_data.values())
        result[ym] = {}
        for dim, ym_dict in month_data.items():
            val = ym_dict.get(ym, 0)
            result[ym][dim] = round(val / total, 4) if total > 0 else 0.0
    return result


# ═══════════════════════════════════════════════════════════
#  모델별 마켓 성장/하락 계산
# ═══════════════════════════════════════════════════════════
def calc_model_mkt_growth(mkt_model_month_data, prev_month, curr_month):
    """
    mkt_model_month_data: {모델명: {마켓: {ym: rev}}}
    prev_month, curr_month: "2026-03", "2026-04" 형태
    반환: {모델명: {best3: [...], worst3: [...]}}
    """
    result = {}
    for model, mkt_dict in mkt_model_month_data.items():
        entries = []
        for market, ym_dict in mkt_dict.items():
            prev_rev = ym_dict.get(prev_month, 0)
            curr_rev = ym_dict.get(curr_month, 0)
            if prev_rev <= 0 and curr_rev <= 0:
                continue
            if prev_rev > 0:
                rate = round((curr_rev - prev_rev) / prev_rev, 4)
            else:
                rate = 1.0 if curr_rev > 0 else 0.0
            entries.append({"market": market, "prevRev": prev_rev, "currRev": curr_rev, "growthRate": rate})

        entries.sort(key=lambda x: x["growthRate"], reverse=True)
        result[model] = {
            "best3":  entries[:3],
            "worst3": list(reversed(entries[-3:])) if entries else [],
        }
    return result


# ═══════════════════════════════════════════════════════════
#  day → month / week 집계 헬퍼
# ═══════════════════════════════════════════════════════════
def _day_to_month(day_dict):
    """
    {날짜str: value} → {ym_str: value}
    예) {"2026-01-05": 100, "2026-01-10": 200} → {"2026-01": 300}
    """
    month_dict = defaultdict(int)
    for ds, val in day_dict.items():
        ym = str(ds)[:7]
        month_dict[ym] += val
    return dict(month_dict)


def _day_to_week(day_dict, weeks):
    """
    {날짜str: value} → {주차str: value} (주차>0 인 날만)
    """
    wk_dict = defaultdict(int)
    for ds, val in day_dict.items():
        try:
            d = date.fromisoformat(str(ds))
        except Exception:
            continue
        wk = weeks.get(d, 0)
        if wk > 0:
            wk_dict[str(wk)] += val
    return dict(wk_dict)


def _build_month_from_day(dim_day_dict):
    """
    {dim: {날짜str: value}} → {dim: {ym: value}}
    """
    return {dim: _day_to_month(dd) for dim, dd in dim_day_dict.items()}


def _build_week_from_day(dim_day_dict, weeks):
    """
    {dim: {날짜str: value}} → {dim: {wk_str: value}}
    """
    return {dim: _day_to_week(dd, weeks) for dim, dd in dim_day_dict.items()}


# ═══════════════════════════════════════════════════════════
#  반품율 계산
# ═══════════════════════════════════════════════════════════
def _calc_return_rates(agg, mkt_day_rev, mkt_day_qty):
    """
    반환: (returnRateMkt, returnQtyMkt, returnRevMkt, returnRateMktMonth)
    """
    ret_rev   = agg.get("ret_mkt_rev", {})
    ret_qty   = agg.get("ret_mkt_qty", {})
    ret_month = agg.get("ret_mkt_month", {})

    return_rate_mkt  = {}
    return_qty_mkt   = {}
    return_rev_mkt   = {}
    for mkt in MARKET_ORDER:
        normal_rev = sum(mkt_day_rev.get(mkt, {}).values())
        ret_r      = ret_rev.get(mkt, 0)
        total      = normal_rev + ret_r
        return_rate_mkt[mkt] = round(ret_r / total, 4) if total > 0 else 0.0
        return_qty_mkt[mkt]  = ret_qty.get(mkt, 0)
        return_rev_mkt[mkt]  = ret_r

    # 월별 반품율
    return_rate_mkt_month = {}
    all_yms = set()
    for ym_dict in ret_month.values():
        all_yms.update(ym_dict.keys())
    for ym in sorted(all_yms):
        return_rate_mkt_month[ym] = {}
        for mkt in MARKET_ORDER:
            ret_r      = ret_month.get(mkt, {}).get(ym, 0)
            normal_rev = sum(v for ds, v in mkt_day_rev.get(mkt, {}).items() if str(ds)[:7] == ym)
            total      = normal_rev + ret_r
            return_rate_mkt_month[ym][mkt] = round(ret_r / total, 4) if total > 0 else 0.0

    return return_rate_mkt, return_qty_mkt, return_rev_mkt, return_rate_mkt_month


# ═══════════════════════════════════════════════════════════
#  mkt_model_rev → {모델: {마켓: {ym: rev}}} 변환
# ═══════════════════════════════════════════════════════════
def _invert_mkt_model_to_model_mkt_month(mkt_model_rev):
    """
    {마켓: {모델: {날짜: rev}}} → {모델: {마켓: {ym: rev}}}
    """
    model_mkt_month = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for mkt, model_dict in mkt_model_rev.items():
        for model, day_dict in model_dict.items():
            for ds, rev in day_dict.items():
                ym = str(ds)[:7]
                model_mkt_month[model][mkt][ym] += rev
    return {m: {mk: dict(ym_dict) for mk, ym_dict in v.items()} for m, v in model_mkt_month.items()}


# ═══════════════════════════════════════════════════════════
#  mktCatShareMonth 계산
# ═══════════════════════════════════════════════════════════
def _calc_mkt_cat_share_month(mkt_model_rev, sku_map):
    """
    {ym: {마켓: {카테고리: 비중}}} 계산
    mkt_model_rev: {마켓: {model_group: {date_str: rev}}}
    sku_map: {sku_base: {"model_group": ..., "category": ...}}
    """
    # model_group → category 역매핑
    model_to_cat = {}
    for info in sku_map.values():
        mg = info.get("model_group", "")
        if mg and mg not in model_to_cat:
            model_to_cat[mg] = info.get("category", "액세서리")

    # mkt×cat×ym 매출 누적
    accum = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for market, model_dict in mkt_model_rev.items():
        for model_grp, day_dict in model_dict.items():
            cat = model_to_cat.get(model_grp) or get_category(model_grp, {})
            for ds, rev in day_dict.items():
                ym = ds[:7]
                accum[ym][market][cat] += rev

    # 비중 계산 (각 ym×market 내 합계 대비)
    result = {}
    for ym, mkt_dict in accum.items():
        result[ym] = {}
        for market, cat_dict in mkt_dict.items():
            total = sum(cat_dict.values())
            if total > 0:
                result[ym][market] = {cat: round(rev / total, 4) for cat, rev in cat_dict.items()}
            else:
                result[ym][market] = {}
    return result


# ═══════════════════════════════════════════════════════════
#  JSON 직렬화 안전 변환
# ═══════════════════════════════════════════════════════════
def _safe(obj):
    """float nan/inf → 0, numpy 타입 → Python 기본 타입으로 변환"""
    import math
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return 0
        return round(obj, 4)
    if isinstance(obj, dict):
        return {str(k): _safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_safe(v) for v in obj]
    # numpy 정수형 처리
    try:
        import numpy as np
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(round(obj, 4))
    except ImportError:
        pass
    return obj


# ═══════════════════════════════════════════════════════════
#  전체 dash_data.json 생성 함수
# ═══════════════════════════════════════════════════════════
def generate_dash_data():
    """
    전체 데이터를 처리하여 dash_data.json을 생성합니다.
    """
    print("=" * 55)
    print("  JBL 판매 대시보드 데이터 엔진")
    print(f"  실행 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 55)

    # ── 1. 마스터 데이터 로드 ──────────────────────────────
    print("\n[1/7] 마스터 데이터 로드 중...")
    df = load_master_data()
    if df.empty:
        print("  [오류] 데이터 없음 — 종료")
        return None

    # ── 2. 보고 연월 감지 ──────────────────────────────────
    print("[2/7] 보고 연월 감지 중...")
    report_year, report_month = detect_report_year_month(df)
    print(f"  → 보고 기준: {report_year}년 {report_month}월")

    # ── 3. 주차 정보 계산 ──────────────────────────────────
    print("[3/7] 주차 정보 계산 중...")
    report_dates, weeks, week_groups, week_labels, max_week = get_week_info(report_year, report_month)
    print(f"  → 주차 수: {max_week}주, 날짜 범위: {report_dates[0]} ~ {report_dates[-1]}")

    # ── 4. SKU 맵 구축 ────────────────────────────────────
    print("[4/7] SKU 맵 구축 중...")
    sku_map = _build_full_sku_map()
    print(f"  → {len(sku_map)}개 SKU 등록")

    # ── 5. 데이터 집계 ────────────────────────────────────
    print("[5/7] 데이터 집계 중...")
    agg = _aggregate_data(df, sku_map, report_dates)
    mkt_day_rev   = agg["mkt_day_rev"]
    mkt_day_qty   = agg["mkt_day_qty"]
    cat_day_rev   = agg["cat_day_rev"]
    cat_day_qty   = agg["cat_day_qty"]
    model_day_rev = agg["model_day_rev"]
    model_day_qty = agg["model_day_qty"]

    # ── 6. 파생 지표 계산 ─────────────────────────────────
    print("[6/7] 파생 지표 계산 중...")

    # 월별 집계
    mkt_month_rev   = _build_month_from_day(mkt_day_rev)
    mkt_month_qty   = _build_month_from_day(mkt_day_qty)
    cat_month_rev   = _build_month_from_day(cat_day_rev)
    cat_month_qty   = _build_month_from_day(cat_day_qty)
    model_month_rev = _build_month_from_day(model_day_rev)
    model_month_qty = _build_month_from_day(model_day_qty)

    # 주차별 집계
    mkt_wk_rev   = _build_week_from_day(mkt_day_rev, weeks)
    mkt_wk_qty   = _build_week_from_day(mkt_day_qty, weeks)
    cat_wk_rev   = _build_week_from_day(cat_day_rev, weeks)
    cat_wk_qty   = _build_week_from_day(cat_day_qty, weeks)
    model_wk_rev = _build_week_from_day(model_day_rev, weeks)
    model_wk_qty = _build_week_from_day(model_day_qty, weeks)

    # MoM
    mkt_mom   = calc_mom(mkt_month_rev)
    cat_mom   = calc_mom(cat_month_rev)
    model_mom = calc_mom(model_month_rev)

    # WoW
    mkt_wow   = calc_wow(mkt_wk_rev)
    cat_wow   = calc_wow(cat_wk_rev)
    model_wow = calc_wow(model_wk_rev)

    # TOP3
    top3_model_month = calc_top3(model_month_rev, model_month_qty, top_n=3)
    top3_model_week  = calc_top3_week(model_wk_rev, model_wk_qty, top_n=3)
    top3_cat_month   = calc_top3(cat_month_rev, cat_month_qty, top_n=3)
    top3_cat_week    = calc_top3_week(cat_wk_rev, cat_wk_qty, top_n=3)

    # 성장/하락
    growth_model_month, decline_model_month = calc_growth_decline(model_month_rev)
    growth_model_week,  decline_model_week  = calc_growth_decline(model_wk_rev)
    growth_cat_month,   decline_cat_month   = calc_growth_decline(cat_month_rev)

    # 비중
    cat_share_month = calc_share(cat_month_rev)
    mkt_share_month = calc_share(mkt_month_rev)

    # 모델-마켓 성장
    model_mkt_month = _invert_mkt_model_to_model_mkt_month(agg["mkt_model_rev"])
    # 최신 2개월 추출
    all_yms = sorted(set(ym for v in mkt_month_rev.values() for ym in v.keys()))
    prev_ym = all_yms[-2] if len(all_yms) >= 2 else None
    curr_ym = all_yms[-1] if all_yms else None
    model_mkt_growth = {}
    if prev_ym and curr_ym:
        model_mkt_growth = calc_model_mkt_growth(model_mkt_month, prev_ym, curr_ym)

    # 반품율
    ret_rate, ret_qty, ret_rev, ret_rate_month = _calc_return_rates(agg, mkt_day_rev, mkt_day_qty)

    # 월별 목표
    monthly_targets = get_monthly_targets()

    # dataMonths: 데이터가 있는 연월 목록
    data_months = sorted(set(ym for v in mkt_month_rev.values() for ym in v.keys()))

    # ── 7. JSON 조립 ──────────────────────────────────────
    print("[7/7] JSON 조립 및 저장 중...")

    # TOP3 key 이름 정리 (model/cat 구분)
    def _rename_top3_model(top3_dict):
        result = {}
        for k, entries in top3_dict.items():
            result[k] = [{"rank": e["rank"], "model": e["name"], "rev": e["rev"], "qty": e["qty"]} for e in entries]
        return result

    def _rename_top3_cat(top3_dict):
        result = {}
        for k, entries in top3_dict.items():
            result[k] = [{"rank": e["rank"], "cat": e["name"], "rev": e["rev"], "qty": e["qty"]} for e in entries]
        return result

    def _rename_growth(gd):
        """name → model"""
        result = {}
        for k, entries in gd.items():
            result[k] = [{"model": e["name"], "prevRev": e["prevRev"], "currRev": e["currRev"], "growthRate": e["growthRate"]} for e in entries]
        return result

    def _rename_growth_cat(gd):
        """name → cat"""
        result = {}
        for k, entries in gd.items():
            result[k] = [{"cat": e["name"], "prevRev": e["prevRev"], "currRev": e["currRev"], "growthRate": e["growthRate"]} for e in entries]
        return result

    dash_data = {
        # 메타
        "updated":      datetime.now().strftime("%Y-%m-%d %H:%M"),
        "period":       f"{data_months[0]} ~ {data_months[-1]}" if data_months else "",
        "reportYear":   report_year,
        "reportMonth":  report_month,
        "dataMonths":   data_months,

        # 차원 목록
        "markets":      MARKET_ORDER,
        "categories":   CATEGORY_ORDER,
        "modelList":    sorted(model_day_rev.keys()),
        "models":       sorted(model_day_rev.keys()),  # 호환성
        "dates":        [d.isoformat() for d in report_dates],

        # 색상
        "mktColors":    {m: MKT_COLORS.get(m, "#607D8B") for m in MARKET_ORDER},
        "catColors":    {c: CAT_COLORS.get(c, "#607D8B") for c in CATEGORY_ORDER},

        # 목표
        "monthlyTargets": monthly_targets,
        "fullMonthlyTargets": monthly_targets,

        # 주차 정보
        "weekLabels":   week_labels,
        "weekCount":    max_week,
        "weekDates":    {str(wk): [d.isoformat() for d in sorted(wd)] for wk, wd in week_groups.items()},

        # 일별
        "mktDayRev":    mkt_day_rev,
        "mktDayQty":    mkt_day_qty,
        "catDayRev":    cat_day_rev,
        "catDayQty":    cat_day_qty,
        # 차트 호환 (동일 데이터, 필드명만 다름)
        "mktDayRevChart": mkt_day_rev,
        "catDayRevChart": cat_day_rev,
        "mktDayQtyChart": mkt_day_qty,
        "catDayQtyChart": cat_day_qty,
        "chartDates":     sorted(set(d for dd in mkt_day_rev.values() for d in dd.keys())),

        # 월별
        "mktMonthRev":  mkt_month_rev,
        "mktMonthQty":  mkt_month_qty,
        "catMonthRev":  cat_month_rev,
        "catMonthQty":  cat_month_qty,
        "modelMonthRev": model_month_rev,
        "modelMonthQty": model_month_qty,

        # 주차별
        "mktWkRev":     mkt_wk_rev,
        "mktWkQty":     mkt_wk_qty,
        "catWkRev":     cat_wk_rev,
        "catWkQty":     cat_wk_qty,
        "modelWkRev":   model_wk_rev,
        "modelWkQty":   model_wk_qty,

        # MoM / WoW
        "mktMoM":       mkt_mom,
        "catMoM":       cat_mom,
        "modelMoM":     model_mom,
        "mktWoW":       mkt_wow,
        "catWoW":       cat_wow,
        "modelWoW":     model_wow,

        # TOP3
        "top3ModelMonth": _rename_top3_model(top3_model_month),
        "top3ModelWeek":  _rename_top3_model(top3_model_week),
        "top3CatMonth":   _rename_top3_cat(top3_cat_month),
        "top3CatWeek":    _rename_top3_cat(top3_cat_week),

        # 성장/하락
        "growthModelMonth":  _rename_growth(growth_model_month),
        "declineModelMonth": _rename_growth(decline_model_month),
        "growthModelWeek":   _rename_growth(growth_model_week),
        "declineModelWeek":  _rename_growth(decline_model_week),
        "growthCatMonth":    _rename_growth_cat(growth_cat_month),
        "declineCatMonth":   _rename_growth_cat(decline_cat_month),

        # 비중
        "catShareMonth": cat_share_month,
        "mktShareMonth": mkt_share_month,
        "mktCatShareMonth": _calc_mkt_cat_share_month(agg["mkt_model_rev"], sku_map),

        # 모델-마켓 성장
        "modelMktGrowth": model_mkt_growth,

        # 반품
        "returnRateMkt":       ret_rate,
        "returnQtyMkt":        ret_qty,
        "returnRevMkt":        ret_rev,
        "returnRateMktMonth":  ret_rate_month,

        # 기존 호환 필드 (streamlit_app.py 사용)
        "targets":        {m: {str(wk): 0 for wk in range(1, max_week + 1)} for m in MARKET_ORDER},
        "totalDayRev":    {ds: sum(mkt_day_rev.get(m, {}).get(ds, 0) for m in MARKET_ORDER)
                           for ds in sorted(set(d for dd in mkt_day_rev.values() for d in dd.keys()))},
        "totalDayQty":    {ds: sum(mkt_day_qty.get(m, {}).get(ds, 0) for m in MARKET_ORDER)
                           for ds in sorted(set(d for dd in mkt_day_qty.values() for d in dd.keys()))},
        "totalWkRev":     {str(wk): sum(mkt_wk_rev.get(m, {}).get(str(wk), 0) for m in MARKET_ORDER)
                           for wk in range(1, max_week + 1)},
        "totalWkQty":     {str(wk): sum(mkt_wk_qty.get(m, {}).get(str(wk), 0) for m in MARKET_ORDER)
                           for wk in range(1, max_week + 1)},
        "modelQtyTargets": {},
        "skus": [],
    }

    # ── JSON 직렬화 안전 처리 후 저장 ─────────────────────
    safe_data = _safe(dash_data)

    try:
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(safe_data, f, ensure_ascii=False, indent=2)
        size_kb = os.path.getsize(OUTPUT_JSON) / 1024
        print(f"\n  [완료] dash_data.json 저장: {size_kb:.1f} KB")
        print(f"  마켓: {list(mkt_day_rev.keys())}")
        print(f"  카테고리: {list(cat_day_rev.keys())}")
        print(f"  모델 수: {len(model_day_rev)}")
        print(f"  데이터 기간: {data_months[0] if data_months else '-'} ~ {data_months[-1] if data_months else '-'}")
    except Exception as e:
        print(f"  [오류] JSON 저장 실패: {e}")
        return None

    return safe_data


# ═══════════════════════════════════════════════════════════
#  기존 대시보드 호환: dash_data.json을 dict로 반환
# ═══════════════════════════════════════════════════════════
def get_dash_data():
    """
    dash_data.json을 읽어 dict로 반환.
    파일이 없으면 generate_dash_data()를 실행하여 생성 후 반환.
    """
    if os.path.exists(OUTPUT_JSON):
        try:
            with open(OUTPUT_JSON, encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"  [경고] dash_data.json 읽기 실패: {e}")

    # 파일 없으면 생성
    return generate_dash_data()


# ═══════════════════════════════════════════════════════════
#  직접 실행
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("데이터 처리 중...")
    result = generate_dash_data()
    if result:
        print("완료! dash_data.json 생성됨")
    else:
        print("실패! 오류 내용을 확인하세요.")
